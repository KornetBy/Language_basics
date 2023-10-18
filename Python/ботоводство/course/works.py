import openpyxl
from aiogram import Bot, Dispatcher, types
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters import Command
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.types import ParseMode
from aiogram.utils import executor

# Инициализация бота и хранилища состояний
bot = Bot(token="5681742144:AAERT6W3vBBY2D8zyig5BVtnYiuDmfrdSj4")
storage = MemoryStorage()
dp = Dispatcher(bot, storage=storage)


# Класс, описывающий состояния пользователя
class Registration(StatesGroup):
    surname = State()  # Фамилия
    name = State()  # Имя
    patronymic = State()  # Отчество
    birthdate = State()  # Дата рождения
    group = State()  # Номер группы
    phone = State()  # Номер телефона
    hostel = State()  # Необходимость общежития
    months = State()  # Месяцы работы
    specialties = State()  # Специальности на каждый месяц работы
    priority = State()  # Приоритеты специальностей


# Открытие таблицы и выбор листа
wb = openpyxl.load_workbook('data.xlsx')
ws = wb['Лист1']


# Обработчик команды /start
@dp.message_handler(Command('start'))
async def start_cmd_handler(message: types.Message):
    # Приветственное сообщение
    await message.answer('Добро пожаловать в бот для вступления в студотряд! Введите /help для просмотра доступных команд.')


# Обработчик команды /help
@dp.message_handler(Command('help'))
async def help_cmd_handler(message: types.Message):
    # Описание доступных команд
    await message.answer('Доступные команды:\n'
                         '/help - помощь\n'
                         '/registration - регистрация\n'
                         '/delete - удаление заявки\n'
                         '/edit - редактирование заявки\n'
                         '/password - ввод кода для повышения приоритета\n'
                         '/about - информация о проекте и создателях\n')


# Обработчик команды /registration
@dp.message_handler(Command('registration'))
async def registration_cmd_handler(message: types.Message):
    # Сообщение о начале регистрации и запрос фамилии
    await message.answer('Введите вашу фамилию:')
    await Registration.surname.set()


# Обработчик ответа на запрос фамилии
@dp.message_handler(state=Registration.surname)
async def surname_handler(message: types.Message, state: FSMContext):
    # Сохранение фамилии в состоянии пользователя и запрос имени
    async with state.proxy() as data:
        data['surname'] = message.text
    await message.answer('Введите ваше имя:')
    await Registration.name.set()


# Обработчик ответа на запрос имени
@dp.message_handler(state=Registration.name)
async def name_handler(message: types.Message, state:FSMContext):
# Сохранение имени в состоянии пользователя и запрос отчества
    async with state.proxy() as data:
        data['name'] = message.text
    await message.answer('Введите ваше отчество:')
    await Registration.patronymic.set()

#Обработчик ответа на запрос отчества
@dp.message_handler(state=Registration.patronymic)
async def patronymic_handler(message: types.Message, state: FSMContext):
# Сохранение отчества в состоянии пользователя и запрос даты рождения
    async with state.proxy() as data:
        data['patronymic'] = message.text
    await message.answer('Введите вашу дату рождения в формате ДД.ММ.ГГГГ:')
    await Registration.birthdate.set()


#Обработчик ответа на запрос даты рождения
@dp.message_handler(state=Registration.birthdate)
async def birthdate_handler(message: types.Message, state: FSMContext):
# Сохранение даты рождения в состоянии пользователя и запрос номера группы
    async with state.proxy() as data:
        data['birthdate'] = message.text
    await message.answer('Введите номер вашей группы:')
    await Registration.group.set()

#Обработчик ответа на запрос номера группы
@dp.message_handler(state=Registration.group)
async def group_handler(message: types.Message, state: FSMContext):
# Сохранение номера группы в состоянии пользователя и запрос номера телефона
    async with state.proxy() as data:
       data['group'] = message.text
    await message.answer('Введите ваш номер телефона:')
    await Registration.phone.set()

#Обработчик ответа на запрос номера телефона
@dp.message_handler(state=Registration.phone)
async def phone_handler(message: types.Message, state: FSMContext):
# Сохранение номера телефона в состоянии пользователя и запрос необходимости общежития
    async with state.proxy() as data:
        data['phone'] = message.text
    await message.answer('Нужно ли вам общежитие? (да/нет)')
    await Registration.hostel.set()

#Обработчик ответа на запрос необходимости общежития
@dp.message_handler(state=Registration.hostel)
async def hostel_handler(message: types.Message, state: FSMContext):
# Сохранение ответа на вопрос о необходимости общежития в состоянии пользователя и запрос количества месяцев работы
    async with state.proxy() as data:
        data['hostel'] = message.text
    await message.answer('Сколько месяцев вы готовы работать в студотряде? (от 1 до 3)')
    await Registration.months.set()

#Обработчик ответа на запрос месяцев работы
@dp.message_handler(state=Registration.months)
async def months_handler(message: types.Message, state: FSMContext):
# Сохранение количества месяцев работы в состоянии пользователя и запрос специальностей
    async with state.proxy() as data:
      data['months'] = message.text
    await message.answer('Введите специальности на каждый месяц работы через запятую (например, "Автомеханик, Электрик, Сварщик"):')
    await Registration.specialties.set()

#Обработчик ответа на запрос специальностей
@dp.message_handler(state=Registration.specialties)
async def specialties_handler(message: types.Message, state: FSMContext):
# Сохранение специальностей в состоянии пользователя и запрос приоритетов специальностей
    async with state.proxy() as data:
        data['specialties'] = message.text.split(', ')
    await message.answer('Введите приоритеты специальностей в порядке убывания важности (например, "2, 1, 3"):')
    await Registration.priority.set()

#Обработчик ответа на запрос приоритетов специальностей
@dp.message_handler(state=Registration.priority)
async def priority_handler(message: types.Message, state: FSMContext):
    # Сохранение приоритетов и вывод собранной информации
    async with state.proxy() as data:
        data['priority'] = message.text
        full_name = f"{data['surname']} {data['name']} {data['patronymic']}"
        registration_data = [
            ['ФИО', full_name],
            ['Дата рождения', data['birthdate']],
            ['Номер группы', data['group']],
            ['Номер телефона', data['phone']],
            ['Нужно ли общежитие', 'Да' if data['hostel'] else 'Нет'],
            ['Месяцы работы', data['months']],
            ['Специальности на каждый месяц работы', ", ".join(f"{data['specialties']}")],
            ['Приоритеты специальностей', data['priority']]]

    # Сохранение информации в таблицу
    row = ws.max_row + 1
    for row_data in registration_data:
        ws.cell(row=row, column=registration_data.index(row_data)+1, value=row_data[1])
    wb.save('data.xlsx')
    
    # Отправка подтверждающего сообщения с собранной информацией
    message_text = 'Вы успешно зарегистрировались! Ваши данные:\n'
    for data_pair in registration_data:
        message_text += f'{data_pair[0]}: {data_pair[1]}\n'
    await message.answer(message_text, parse_mode=ParseMode.MARKDOWN)

    # Сброс состояния пользователя
    await state.finish()





#Обработчик команды /about
@dp.message_handler(Command('about'))
async def about_cmd_handler(message: types.Message):
# Информация о проекте и создателях
    about_text = 'Этот бот создан для упрощения процесса вступления в студотряд.\n\n'
    'Авторы проекта:\n'
    'Иванов Иван - @ivanov\n'
    'Петров Петр - @petrov'
    await message.answer(about_text)

#Обработчик ошибки, возникающей при некорректном вводе
@dp.message_handler()
async def undefined_cmd_handler(message: types.Message):
    await message.answer('Я не понимаю, что вы хотите. Введите /help для просмотра доступных команд.')

#Обработчик ошибки, возникающей при некорректном вводе при регистрации
@dp.message_handler(state='*', content_types=types.ContentTypes.TEXT)
async def process_message(message: types.Message, state: FSMContext):
    await message.answer('Произошла ошибка при вводе. Попробуйте снова.')
    #await state.finish()

#Функция для сохранения заявки в таблице Excel
def save_registration_to_excel(registration_data):
# Получение номера первой свободной строки в таблице
    row_number = ws.max_row + 1

    # Заполнение ячеек таблицы данными из заявки
    ws.cell(row=row_number, column=1, value=row_number - 1)
    ws.cell(row=row_number, column=2, value=registration_data['surname'])
    ws.cell(row=row_number, column=3, value=registration_data['name'])
    ws.cell(row=row_number, column=4, value=registration_data['patronymic'])
    ws.cell(row=row_number, column=5, value=registration_data['birthdate'])
    ws.cell(row=row_number, column=6, value=registration_data['group'])
    ws.cell(row=row_number, column=7, value=registration_data['phone'])
    ws.cell(row=row_number, column=8, value=registration_data['hostel'])
    ws.cell(row=row_number, column=9, value=registration_data['months'])
    ws.cell(row=row_number, column=10, value=registration_data['specialties'])
    ws.cell(row=row_number, column=11, value=registration_data['priority'])

    # Сохранение изменений в таблице
    wb.save('data.xlsx')

#Функция для удаления заявки из таблицы Excel
def delete_registration_from_excel(registration_id):
# Поиск строки с номером удаляемой заявки
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == registration_id:
        # Удаление строки
            ws.delete_rows(row, amount=1)
        # Сохранение изменений в таблице
        wb.save('data.xlsx')
        return True
    return False

# Функция для редактирования заявки в таблице Excel
async def edit_application(application_id: str, column: str, new_value: str):
    # Получение строки с заявкой по идентификатору
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == application_id:
            # Изменение значения в указанном столбце
            index = ws.cell(row=1, column=column).column
            ws.cell(row=row[0], column=index, value=new_value)
            # Сохранение изменений в таблице
            wb.save('data.xlsx')
            return True
    return False

if __name__== '__main__':
    executor.start_polling(dp)